import io
from datetime import datetime
from typing import List, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from app.services.google_sheets import google_sheets_service

class ExcelExportService:
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_individual_timesheet(self, timesheet_data: Dict[str, Any], user_info: Dict[str, Any]) -> bytes:
        """Export individual timesheet to Excel"""
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Timesheet"
        
        # Add header information
        ws['A1'] = "Simple Timesheet Export"
        ws['A1'].font = Font(size=16, bold=True)
        
        ws['A3'] = "Employee:"
        ws['B3'] = user_info.get('full_name', 'Unknown')
        ws['A4'] = "Email:"
        ws['B4'] = user_info.get('email', 'Unknown')
        ws['A5'] = "Period:"
        ws['B5'] = f"{timesheet_data.get('period_start', '')} to {timesheet_data.get('period_end', '')}"
        ws['A6'] = "Status:"
        ws['B6'] = timesheet_data.get('status', 'Unknown').upper()
        ws['A7'] = "Total Hours:"
        ws['B7'] = timesheet_data.get('total_hours', 0)
        
        # Get timesheet data from Google Sheets
        sheet_data = []
        if timesheet_data.get('google_sheet_url'):
            try:
                sheet_data = google_sheets_service.get_timesheet_data(timesheet_data['google_sheet_url'])
            except Exception as e:
                print(f"Error fetching sheet data: {e}")
                sheet_data = []
        
        # Add timesheet data table
        if sheet_data:
            start_row = 10
            
            # Headers
            headers = ["Date", "Start Time", "End Time", "Break (mins)", "Total Hours", "Project", "Description", "Status"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.border
            
            # Data rows
            for row_idx, entry in enumerate(sheet_data, start_row + 1):
                ws.cell(row=row_idx, column=1, value=entry.get('date', '')).border = self.border
                ws.cell(row=row_idx, column=2, value=entry.get('start_time', '')).border = self.border
                ws.cell(row=row_idx, column=3, value=entry.get('end_time', '')).border = self.border
                ws.cell(row=row_idx, column=4, value=entry.get('break_duration', 0)).border = self.border
                ws.cell(row=row_idx, column=5, value=entry.get('total_hours', 0)).border = self.border
                ws.cell(row=row_idx, column=6, value=entry.get('project', '')).border = self.border
                ws.cell(row=row_idx, column=7, value=entry.get('task_description', '')).border = self.border
                ws.cell(row=row_idx, column=8, value=entry.get('status', '')).border = self.border
        else:
            ws['A10'] = "No timesheet data available or unable to fetch from Google Sheets."
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 12
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def export_team_timesheets(self, timesheets_data: List[Dict[str, Any]], supervisor_info: Dict[str, Any]) -> bytes:
        """Export multiple team timesheets to Excel"""
        wb = Workbook()
        
        # Summary sheet
        summary_ws = wb.active
        summary_ws.title = "Summary"
        
        # Add header
        summary_ws['A1'] = "Team Timesheet Summary"
        summary_ws['A1'].font = Font(size=16, bold=True)
        
        summary_ws['A3'] = "Supervisor:"
        summary_ws['B3'] = supervisor_info.get('full_name', 'Unknown')
        summary_ws['A4'] = "Generated:"
        summary_ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        summary_ws['A5'] = "Total Timesheets:"
        summary_ws['B5'] = len(timesheets_data)
        
        # Summary table headers
        start_row = 8
        headers = ["Staff Member", "Period", "Status", "Total Hours", "Submitted Date", "Reviewed Date"]
        for col, header in enumerate(headers, 1):
            cell = summary_ws.cell(row=start_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Summary data
        for row_idx, timesheet in enumerate(timesheets_data, start_row + 1):
            summary_ws.cell(row=row_idx, column=1, value=timesheet.get('staff_name', 'Unknown')).border = self.border
            summary_ws.cell(row=row_idx, column=2, value=timesheet.get('period', '')).border = self.border
            summary_ws.cell(row=row_idx, column=3, value=timesheet.get('status', '').upper()).border = self.border
            summary_ws.cell(row=row_idx, column=4, value=timesheet.get('total_hours', 0)).border = self.border
            summary_ws.cell(row=row_idx, column=5, value=timesheet.get('submitted_at', '')).border = self.border
            summary_ws.cell(row=row_idx, column=6, value=timesheet.get('reviewed_at', '')).border = self.border
        
        # Adjust column widths for summary
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            summary_ws.column_dimensions[col].width = 18
        
        # Individual sheets for each timesheet (optional - can be enabled if needed)
        # for timesheet in timesheets_data:
        #     if timesheet.get('google_sheet_url'):
        #         try:
        #             sheet_data = google_sheets_service.get_timesheet_data(timesheet['google_sheet_url'])
        #             if sheet_data:
        #                 self._add_individual_sheet(wb, timesheet, sheet_data)
        #         except Exception as e:
        #             print(f"Error adding individual sheet: {e}")
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def _add_individual_sheet(self, wb: Workbook, timesheet: Dict[str, Any], sheet_data: List[Dict[str, Any]]):
        """Add individual timesheet as a separate sheet"""
        staff_name = timesheet.get('staff_name', 'Unknown').replace(' ', '_')
        period = timesheet.get('period', '').replace(' ', '_')
        sheet_name = f"{staff_name}_{period}"[:31]  # Excel sheet name limit
        
        ws = wb.create_sheet(title=sheet_name)
        
        # Headers
        headers = ["Date", "Start Time", "End Time", "Break (mins)", "Total Hours", "Project", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Data
        for row_idx, entry in enumerate(sheet_data, 2):
            ws.cell(row=row_idx, column=1, value=entry.get('date', '')).border = self.border
            ws.cell(row=row_idx, column=2, value=entry.get('start_time', '')).border = self.border
            ws.cell(row=row_idx, column=3, value=entry.get('end_time', '')).border = self.border
            ws.cell(row=row_idx, column=4, value=entry.get('break_duration', 0)).border = self.border
            ws.cell(row=row_idx, column=5, value=entry.get('total_hours', 0)).border = self.border
            ws.cell(row=row_idx, column=6, value=entry.get('project', '')).border = self.border
            ws.cell(row=row_idx, column=7, value=entry.get('task_description', '')).border = self.border
        
        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 30

# Global instance
excel_export_service = ExcelExportService()